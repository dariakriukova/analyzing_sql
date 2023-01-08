from datetime import datetime, date
from decimal import Decimal
from os.path import exists as file_exists
from time import strftime, localtime
from typing import Union, Any, TypeVar, Generic, Callable, Iterable

import click
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from matplotlib.ticker import MaxNLocator
from numpy import ndarray
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill, Fill
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import desc
from sqlalchemy.orm import Session, joinedload

from orm import IVMeasurement, CVMeasurement, Wafer, Chip, ChipState
from utils import (
    logger,
    flatten_options,
    iv_thresholds,
    cv_thresholds,
    IV_VOLTAGE_PRESETS,
    VoltagesOption,
    EntityChoice
)

date_formats = ['%Y-%m-%dT%H:%M:%S', '%Y-%m-%d']
date_formats_help = f"Supported formats are: {', '.join((strftime(f) for f in date_formats))}."


@click.command(name='iv', help="Make summary (png and xlsx) for IV measurements' data.")
@click.pass_context
@click.option("-t", "--chips-type", help="Type of the chips to analyze.")
@click.option("-w", "--wafer", "wafer_name", prompt=f"Wafer name", help="Wafer name.")
@click.option("-o", "--output", "file_name",
              default=lambda: f"summary-iv-{strftime('%y%m%d-%H%M%S')}",
              help="Output file names without extension.", show_default="summary-iv-{datetime}")
@click.option("-s", "--chip-state", "chip_state_ids", help="State of the chips to analyze.",
              default=['ALL'], show_default=True, multiple=True, callback=flatten_options)
@click.option("--outliers-coefficient", default=2.0, show_default=True,
              help="Standard deviation multiplier to detect outlier measurements.", type=float)
@click.option("--before", type=click.DateTime(formats=date_formats),
              help=f"Include measurements before (exclusive) provided date and time. {date_formats_help}")
@click.option("--after", type=click.DateTime(formats=date_formats),
              help=f"Include measurements after (inclusive) provided date and time. {date_formats_help}")
@click.option("--voltages", "voltages", default=IV_VOLTAGE_PRESETS['sm'],
              cls=VoltagesOption, presets=IV_VOLTAGE_PRESETS,
              help="List of voltages to include in summary.")
def summary_iv(ctx: click.Context, chips_type: Union[str, None], wafer_name: str, file_name: str,
               chip_state_ids: tuple[str], outliers_coefficient: float,
               before: Union[datetime, None],
               after: Union[datetime, None],
               voltages: Iterable[Decimal]):
    session: Session = ctx.obj['session']
    if ctx.obj['default_wafer'].name != wafer_name:
        wafer = session.query(Wafer).filter(Wafer.name == wafer_name).first()
    else:
        wafer = ctx.obj['default_wafer']
    query = session.query(IVMeasurement) \
        .filter(IVMeasurement.chip.has(Chip.wafer.__eq__(wafer))) \
        .options(joinedload(IVMeasurement.chip))

    if chips_type is not None:
        query = query.filter(IVMeasurement.chip.has(Chip.type.__eq__(chips_type)))
    else:
        logger.info('Chips type (-t or --chips-type) is not specified. Analyzing all chip types.')

    query = query.filter(IVMeasurement.chip_state_id.in_(chip_state_ids))

    if before is not None or after is not None:
        after = after if after is not None else date.min
        before = before if before is not None else date.max
        query = query.filter(IVMeasurement.datetime.between(after, before))

    measurements = query.all()

    if not measurements:
        logger.warn('No measurements found.')
        return

    sheets_data = get_sheets_data(measurements)
    value_extractor = lambda m: m.anode_current_corrected or m.anode_current
    fig, axes = plot_data(measurements, voltages, outliers_coefficient,
                          value_extractor)
    for [ax, _] in axes:
        ax.set_xlabel("Anode current [pA]")

    png_file_name = file_name + '.png'
    check_file_exists(png_file_name)
    fig.savefig(png_file_name, dpi=300)
    logger.info(f'Summary data is plotted to {png_file_name}')

    exel_file_name = file_name + '.xlsx'
    check_file_exists(exel_file_name)
    info = get_info(ctx, wafer=wafer, chip_state_ids=chip_state_ids, measurements=measurements)
    save_iv_summary_to_excel(sheets_data, info, exel_file_name, voltages)

    logger.info(f'Summary data is saved to {exel_file_name}')


@click.command(name='cv', help="Make summary_group (png and xlsx) for CV measurements' data.")
@click.pass_context
@click.option("-t", "--chips-type", help="Type of the chips to analyze.")
@click.option("-w", "--wafer", "wafer_name", prompt=f"Wafer name", help="Wafer name.")
@click.option("-o", "--output", "file_name",
              default=lambda: f"summary-cv-{strftime('%y%m%d-%H%M%S')}",
              help="Output file names without extension.", show_default="summary-cv-{datetime}")
@click.option("-s", "--chip-state", "chip_state_ids", help="State of the chips to analyze.",
              default=['ALL'], show_default=True, multiple=True, callback=flatten_options)
@click.option("--outliers-coefficient", default=2.0, show_default=True,
              help="Standard deviation multiplier to detect outlier measurements.", type=float)
@click.option("--before", type=click.DateTime(formats=date_formats),
              help=f"Include measurements before (exclusive) provided date and time. {date_formats_help}")
@click.option("--after", type=click.DateTime(formats=date_formats),
              help=f"Include measurements after (inclusive) provided date and time. {date_formats_help}")
@click.option("--voltages", "voltages", default=["-5", "0", "-35"], multiple=True,
              show_default=True, callback=flatten_options,
              help="List of voltages to include in summary.")
def summary_cv(ctx: click.Context, chips_type: Union[str, None], wafer_name: str, file_name: str,
               chip_state_ids: list[str], outliers_coefficient: float,
               before: Union[datetime, None],
               after: Union[datetime, None],
               voltages: set[str]):
    session: Session = ctx.obj['session']
    if ctx.obj['default_wafer'].name != wafer_name:
        wafer = session.query(Wafer).filter(Wafer.name == wafer_name).first()
    else:
        wafer = ctx.obj['default_wafer']
    query = session.query(CVMeasurement) \
        .filter(CVMeasurement.chip.has(Chip.wafer.__eq__(wafer))) \
        .options(joinedload(CVMeasurement.chip))

    if chips_type is not None:
        query = query.filter(CVMeasurement.chip.has(Chip.type.__eq__(chips_type)))
    else:
        logger.info('Chips type (-t or --chips-type) is not specified. Analyzing all chip types.')

    query = query.filter(CVMeasurement.chip_state_id.in_(chip_state_ids))

    if before is not None or after is not None:
        after = after if after is not None else date.min
        before = before if before is not None else date.max
        query = query.filter(CVMeasurement.datetime.between(after, before))

    measurements = query.all()

    if not measurements:
        logger.warn('No measurements found.')
        return

    sheets_data = get_sheets_cv_data(measurements)
    value_extractor = lambda m: m.capacitance
    voltages = sorted(Decimal(v) for v in voltages)
    fig, axes = plot_data(measurements, voltages, outliers_coefficient,
                          value_extractor)
    for [ax, _] in axes:
        ax.set_xlabel("Capacitance [pF]")

    png_file_name = file_name + '.png'
    check_file_exists(png_file_name)
    fig.savefig(png_file_name, dpi=300)
    logger.info(f'Summary data is plotted to {png_file_name}')

    exel_file_name = file_name + '.xlsx'
    check_file_exists(exel_file_name)
    info = get_info(ctx, wafer=wafer, chip_state_ids=chip_state_ids, measurements=measurements)
    save_cv_summary_to_excel(sheets_data, info, exel_file_name, voltages)

    logger.info(f'Summary data is saved to {exel_file_name}')


@click.group(name='summary', help="Group of command to analyze and summaryze the data",
             commands=[summary_iv, summary_cv])
@click.pass_context
def summary_group(ctx: click.Context):
    session: Session = ctx.obj['session']
    active_command = summary_group.commands[ctx.invoked_subcommand]
    last_wafer = session.query(Wafer).order_by(desc(Wafer.record_created_at)).first()
    default_wafer_name = last_wafer.name
    wafer_option = next((o for o in active_command.params if o.name == 'wafer_name'))
    wafer_option.default = default_wafer_name
    ctx.obj['default_wafer'] = last_wafer

    chip_states = session.query(ChipState).order_by(ChipState.id).all()
    ctx.obj['chip_states'] = chip_states
    chip_state_option = next(
        (o for o in active_command.params if o.name == 'chip_state_ids'))
    chip_state_option.type = EntityChoice(choices=chip_states, multiple=chip_state_option.multiple)


def save_iv_summary_to_excel(sheets_data: dict, info: pd.Series, file_name: str,
                             voltages: Iterable[Decimal]):
    summary_df = get_slice_by_voltages(sheets_data['anode'], voltages)
    rules = {
        'lessThan': PatternFill(bgColor='ee9090', fill_type='solid'),
        'greaterThanOrEqual': PatternFill(bgColor='90ee90', fill_type='solid')
    }

    with pd.ExcelWriter(file_name) as writer:
        summary_df.rename(columns=float).to_excel(writer, sheet_name='Summary')
        apply_conditional_formatting(writer.book["Summary"], sheets_data['chip_types'], rules,
                                     iv_thresholds)

        sheets_data['anode'].rename(columns=float).to_excel(writer, sheet_name='I1 anode')
        sheets_data['cathode'].rename(columns=float).to_excel(writer, sheet_name='I3 cathode')
        info.to_excel(writer, sheet_name='Info')


def save_cv_summary_to_excel(sheets_data: dict, info: pd.Series, file_name: str,
                             voltages: Iterable[Decimal]):
    summary_df = get_slice_by_voltages(sheets_data['capacitance'], voltages)
    rules = {
        'greaterThanOrEqual': PatternFill(bgColor='ee9090', fill_type='solid'),
        'lessThan': PatternFill(bgColor='90ee90', fill_type='solid')
    }

    with pd.ExcelWriter(file_name) as writer:
        summary_df.to_excel(writer, sheet_name='Summary')
        apply_conditional_formatting(writer.book["Summary"], sheets_data['chip_types'], rules,
                                     cv_thresholds)

        sheets_data['capacitance'].rename(columns=float).to_excel(writer, sheet_name='All data')
        info.to_excel(writer, sheet_name='Info')


def get_slice_by_voltages(df: pd.DataFrame, voltages: Iterable[Decimal]) -> pd.DataFrame:
    columns = sorted(voltages)
    slice_df = pd.DataFrame(columns=columns)
    slice_df = pd.concat((slice_df, df[df.columns.intersection(columns)]), copy=False)

    empty_cols = slice_df.isna().all(axis=0)
    if empty_cols.any():
        voltages_option = next(p for p in summary_iv.params if p.name == 'voltages')
        logger.warn(f"""
        The following voltages are not present in the data: {
        [float(col) for col, val in empty_cols.items() if val]
        }.
        Use the {voltages_option.opts[0]} option to select the existing voltages.
        Available voltages: {[float(voltage) for voltage in df.columns]}
        """)
    return slice_df


def apply_conditional_formatting(sheet: Worksheet, chip_types: list[str], rules: dict[str, Fill],
                                 thresholds: dict[str, dict[str, float]]):
    chip_row_index = [(i + 1, cell.value) for i, cell in enumerate(sheet['A']) if cell.value]

    for chip_type in chip_types:
        def is_current_type(chip_name: str) -> bool:
            return chip_name.startswith(chip_type)

        for voltage, threshold in thresholds[chip_type].items():
            voltage = Decimal(voltage)
            try:
                column_cell = next((cell for cell in sheet['1'] if
                                    cell.value is not None and Decimal(str(cell.value)) == voltage))
                first_row_index = next(i for i, v in chip_row_index if is_current_type(v))
                last_row_index = next(i for i, v in reversed(chip_row_index) if is_current_type(v))
            except ValueError:
                continue
            except StopIteration:
                continue

            column_letter = column_cell.column_letter
            cell_range = f'{column_letter}{first_row_index}:{column_letter}{last_row_index}'
            for rule_name, rule in rules.items():
                cells_rule = CellIsRule(operator=rule_name, formula=[threshold], fill=rule)
                sheet.conditional_formatting.add(cell_range, cells_rule)


def get_sheets_data(measurements: list[IVMeasurement]) -> dict[str, Union[pd.DataFrame, Any]]:
    chip_names = sorted({measurement.chip.name for measurement in measurements})
    chip_types = {measurement.chip.type for measurement in measurements}
    voltages = sorted({measurement.voltage_input for measurement in measurements})
    anode_df = pd.DataFrame(dtype='float64', index=chip_names, columns=voltages)
    cathode_df = pd.DataFrame(dtype='float64', index=chip_names, columns=voltages)
    has_uncorrected_current = False
    with click.progressbar(measurements, label='Processing measurements...') as progress:
        for measurement in progress:
            measurement: IVMeasurement
            cell_location = (measurement.chip.name, measurement.voltage_input)
            if measurement.anode_current_corrected is None:
                has_uncorrected_current = True
                anode_df.loc[cell_location] = measurement.anode_current
            else:
                anode_df.loc[cell_location] = measurement.anode_current_corrected
            cathode_df.loc[cell_location] = measurement.cathode_current
    if has_uncorrected_current:
        logger.warning('Some current measurements are not corrected by temperature.')
    return {
        'anode': anode_df,
        'cathode': cathode_df,
        'chip_names': chip_names,
        'chip_types': chip_types,
        'voltages': voltages
    }


def get_sheets_cv_data(measurements: list[CVMeasurement]) -> dict[str, Union[pd.DataFrame, Any]]:
    chip_names = sorted({measurement.chip.name for measurement in measurements})
    chip_types = {measurement.chip.type for measurement in measurements}
    voltages = sorted({measurement.voltage_input for measurement in measurements})
    capacitance_df = pd.DataFrame(dtype='float64', index=chip_names, columns=voltages)
    with click.progressbar(measurements, label='Processing measurements...') as progress:
        for measurement in progress:
            measurement: CVMeasurement
            cell_location = (measurement.chip.name, measurement.voltage_input)
            capacitance_df.loc[cell_location] = measurement.capacitance
    return {
        'capacitance': capacitance_df,
        'chip_names': chip_names,
        'chip_types': chip_types,
        'voltages': voltages
    }


def get_info(ctx: click.Context, wafer: Wafer, chip_state_ids: Iterable[str],
             measurements: list[Union[IVMeasurement, CVMeasurement]]) -> pd.Series:
    format_date = strftime("%A, %d %b %Y", localtime())
    chip_states_str = "; ".join(
        [state.name for state in ctx.obj['chip_states'] if str(state.id) in chip_state_ids])

    first_measurement = min(measurements, key=lambda m: m.datetime)
    last_measurement = max(measurements, key=lambda m: m.datetime)

    return pd.Series({
        'Wafer': wafer.name,
        'Summary generation date': format_date,
        'Chip state': chip_states_str,
        "First measurement date": first_measurement.datetime,
        "Last measurement date": last_measurement.datetime,
    })


def check_file_exists(file_name: str):
    while True:
        if file_exists(file_name):
            ok = click.prompt(f'File {file_name} already exists. Overwrite?', default='y',
                              type=click.Choice(['y', 'n']))
            if ok == 'n':
                file_name = click.prompt(f'Enter new file name')
            else:
                break
        else:
            break


def get_outliers_idx(data: np.ndarray, m: float) -> np.ndarray:
    return np.abs(data - np.nanmedian(data)) > m * np.nanstd(data)


def plot_hist(ax: Axes, data: np.ndarray):
    ax.set_ylabel("Number of chips")
    ax.hist(data * 1e12, bins=15)


T = TypeVar('T')


def plot_heat_map(ax: Axes, measurements: list[Generic[T]], low, high,
                  value_extractor: Callable[[T], float]):
    xs = {measurement.chip.x_coordinate for measurement in measurements}
    ys = {measurement.chip.y_coordinate for measurement in measurements}

    width = max(xs) - min(xs) + 1
    height = max(ys) - min(ys) + 1
    grid = np.full((height, width), np.nan)
    for cell in measurements:
        grid[cell.chip.y_coordinate - min(ys)][cell.chip.x_coordinate - min(xs)] = value_extractor(
            cell)

    X = np.linspace(min(xs) - 0.5, max(xs) + 0.5, width + 1)
    Y = np.linspace(min(ys) - 0.5, max(ys) + 0.5, height + 1)
    mesh = ax.pcolormesh(X, Y, grid, cmap='hot', shading='flat', vmin=low, vmax=high)
    ax.xaxis.set_major_locator(MaxNLocator(integer=True, min_n_ticks=0))
    ax.yaxis.set_major_locator(MaxNLocator(integer=True, min_n_ticks=0))
    ax.set_ylabel("Y coordinate")
    ax.set_xlabel("X coordinate")
    ax.figure.colorbar(mesh, ax=ax)


def plot_data(values: list[Generic[T]], voltages: Iterable[Decimal],
              outliers_coefficient: float, value_extractor: Callable[[T], float]) -> (
        Figure, ndarray[Any, Axes]):
    fig, axes = plt.subplots(nrows=len(voltages), ncols=2,
                             figsize=(10, 5 * len(voltages)),
                             gridspec_kw=dict(left=0.08, right=0.95, bottom=0.05, top=0.95,
                                              wspace=0.3, hspace=0.35))
    axes = axes.reshape(-1, 2)

    for i, voltage in enumerate(sorted(voltages)):
        target_values = [value for value in values if value.voltage_input == voltage]
        if len(target_values) == 0:
            continue
        data = np.array([value_extractor(value) for value in target_values])

        outliers_idx = get_outliers_idx(data, outliers_coefficient)
        if outliers_idx.any():
            outlier_chip_names = (outlier.chip.name for outlier in
                                  np.array(target_values)[outliers_idx])
            logger.warn(
                f'Outliers detected! {", ".join(outlier_chip_names)} are ignored on {voltage}V histogram and heat map color scale')
            data = data[~outliers_idx]

        axes[i][0].set_title(f"{voltage}V")
        plot_hist(axes[i][0], data)

        low, high = data.min(), data.max()
        axes[i][1].set_title(f"{voltage}V")
        plot_heat_map(axes[i][1], target_values, low, high, value_extractor)
    return fig, axes
