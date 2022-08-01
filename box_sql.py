from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import pyvisa
import re
import datetime
from time import sleep
from scipy.optimize import curve_fit
import numpy as np
from get_temperature import get_temperature
import string
import mysql.connector
from mysql.connector import errorcode


TARGET_TEMPERATURE = 25

configs = {
    "VAR1_START": 0.01,
    "VAR1_STEP": 0.001,
    "VAR1_END": -0.01,
    "C_VAR1": 0.1,  # compliance
    "HOLD_TIME": 0,
    "DELAY_TIME": 0,
    "INT_TIME": 'SHORT',
    "INT_TIME_SHORT": '',
    "INT_TIME_MED": '',
    "INT_TIME_LONG": '',
    "VAR1_MODE": 'SINGLE',
    "VAR1_SWEEP_TYPE": 'LIN',
    "VOLT_OUTPUT_SMU1": 'V',
    "VOLT_OUTPUT_SMU2": 'V',
    "VOLT_OUTPUT_SMU3": 'V',
    "VOLT_OUTPUT_SMU4": 'V',
    "FUNCTION_SMU1": 'CONS',
    "FUNCTION_SMU2": 'CONS',
    "FUNCTION_SMU3": 'CONS',
    "FUNCTION_SMU4": 'VAR1',
    "V_NAME_SMU1": 'V1',
    "V_NAME_SMU2": 'V2',
    "V_NAME_SMU3": 'V3',
    "V_NAME_SMU4": 'V4',
    "I_NAME_SMU1": 'I1',
    "I_NAME_SMU2": 'I2',
    "I_NAME_SMU3": 'I3',
    "I_NAME_SMU4": 'I4',
    "RANGE_MODE_SMU1": '',
    "RANGE_MODE_SMU2": '',
    "RANGE_MODE_SMU3": '',
    "RANGE_MODE_SMU4": '',
    "RANGE_VALUE_SMU1": '',
    "RANGE_VALUE_SMU2": '',
    "RANGE_VALUE_SMU3": '',
    "RANGE_VALUE_SMU4": '',
    "DATETIME": datetime.datetime.now(),
    "TEMPERATURE": get_temperature()
}

chip_1 = input('Chip 1:')
chip_2 = input('Chip 2:')
chip_3 = input('Chip 3:')
WAFER = 'BP6'
stage = 'after to-can packaging'

rm = pyvisa.ResourceManager()
instrument = rm.open_resource('GPIB0::15::INSTR')  # the instrument
instrument.baud_rate = 38400
instrument.timeout = 250000  # If error "Timeout is expired" appears, this value can be increased
# print(instrument.query('*IDN?'))  # print the name of instrument


def set_configs(conf):
    instrument.write('*CLS')  # clears the Error Queue
    instrument.write('*RST')  # performs an instrument reset

    instrument.write(":PAGE:MEAS:VAR1:MODE {}".format(conf["VAR1_MODE"]))  # sets the SWEEP MODE of VAR1 for sweep measurement.(Single)
    instrument.write(":PAGE:MEAS:VAR1:SPAC {}".format(conf["VAR1_SWEEP_TYPE"]))  # selects the sweep type of VAR1 as linear staircase
    instrument.write(":PAGE:MEAS:VAR1:START {}".format(conf['VAR1_START']))  # sets the sweep START value of VAR1.
    instrument.write(":PAGE:MEAS:VAR1:STEP {}".format(conf['VAR1_STEP']))  # sets the sweep STEP value of VAR1 for the linear sweep.
    instrument.write(":PAGE:MEAS:VAR1:STOP {}".format(conf['VAR1_END']))  # sets the sweep STOP value of VAR1.
    instrument.write(":PAGE:MEAS:VAR1:COMP {}".format(conf['C_VAR1']))  # sets the COMPLIANCE value of VAR1.

    instrument.write(":PAGE:MEAS:VAR1:PCOM:STATE OFF")  # sets the power compliance of VAR1 to disable

    instrument.write(":PAGE:MEAS:HTIM {}".format(conf['HOLD_TIME']))  # sets the HOLD TIME of sweep measurement
    instrument.write(":PAGE:MEAS:DEL {}".format(conf['DELAY_TIME']))  # sets the DELAY TIME of SMU.

    # Integration time
    instrument.write(":PAGE:MEAS:MSET:ITIM {}".format(conf['INT_TIME']))
    instrument.write(":PAGE:MEAS:MSET:ITIM:SHORT {}".format(conf['INT_TIME_SHORT']))
    instrument.write(":PAGE:MEAS:MSET:ITIM:MED {}".format(conf['INT_TIME_MED']))
    instrument.write(":PAGE:MEAS:MSET:ITIM:LONG {}".format(conf['INT_TIME_LONG']))

    instrument.write(":PAGE:CHAN:SMU1:MODE {}".format(conf['VOLT_OUTPUT_SMU1']))  # sets the voltage output MODE of SMU1
    instrument.write(":PAGE:CHAN:SMU2:MODE {}".format(conf['VOLT_OUTPUT_SMU2']))  # sets the voltage output MODE of SMU2
    instrument.write(":PAGE:CHAN:SMU3:MODE {}".format(conf['VOLT_OUTPUT_SMU3']))  # sets the voltage output MODE of SMU1
    instrument.write(":PAGE:CHAN:SMU4:MODE {}".format(conf['VOLT_OUTPUT_SMU4']))

    instrument.write(":PAGE:CHAN:SMU1:FUNC {}".format(conf['FUNCTION_SMU1']))  # sets the function (FCTN) of SMU1. SMU1 - Emitter
    instrument.write(":PAGE:CHAN:SMU2:FUNC {}".format(conf['FUNCTION_SMU2']))  # sets the function (FCTN) of SMU2 as a constant
    instrument.write(":PAGE:CHAN:SMU3:FUNC {}".format(conf['FUNCTION_SMU3']))  # sets the function (FCTN) of SMU3. SMU3 - Base
    instrument.write(":PAGE:CHAN:SMU4:FUNC {}".format(conf['FUNCTION_SMU4']))
    # Ranging mode
    instrument.write(":PAGE:MEAS:MSET:SMU1:RANG:MODE {}".format(conf['RANGE_MODE_SMU1']))
    instrument.write(":PAGE:MEAS:MSET:SMU2:RANG:MODE {}".format(conf['RANGE_MODE_SMU2']))
    instrument.write(":PAGE:MEAS:MSET:SMU3:RANG:MODE {}".format(conf['RANGE_MODE_SMU3']))
    instrument.write(":PAGE:MEAS:MSET:SMU4:RANG:MODE {}".format(conf['RANGE_MODE_SMU4']))
    instrument.write(":PAGE:MEAS:MSET:SMU1:RANG {}".format(conf['RANGE_VALUE_SMU1']))
    instrument.write(":PAGE:MEAS:MSET:SMU2:RANG {}".format(conf['RANGE_VALUE_SMU2']))
    instrument.write(":PAGE:MEAS:MSET:SMU3:RANG {}".format(conf['RANGE_VALUE_SMU3']))
    instrument.write(":PAGE:MEAS:MSET:SMU4:RANG {}".format(conf['RANGE_VALUE_SMU4']))

    instrument.write(":PAGE:CHAN:SMU1:VNAME '{}'".format(conf['V_NAME_SMU1']))
    instrument.write(":PAGE:CHAN:SMU2:VNAME '{}'".format(conf['V_NAME_SMU2']))
    instrument.write(":PAGE:CHAN:SMU3:VNAME '{}'".format(conf['V_NAME_SMU3']))
    instrument.write(":PAGE:CHAN:SMU4:VNAME '{}'".format(conf['V_NAME_SMU4']))
    instrument.write(":PAGE:CHAN:SMU2:INAME '{}'".format(conf['I_NAME_SMU2']))
    instrument.write(":PAGE:CHAN:SMU3:INAME '{}'".format(conf['I_NAME_SMU3']))
    instrument.write(":PAGE:CHAN:SMU4:INAME '{}'".format(conf['I_NAME_SMU4']))
    instrument.write(":PAGE:CHAN:SMU1:INAME '{}'".format(conf['I_NAME_SMU1']))

    instrument.write(":PAGE:DISP:LIST 'V4','I1', 'I2', 'I3', 'I4'")  # selects the variable names for LIST display
    instrument.write(":PAGE:CHAN:MODE SWEEP")  # sets the sweep measurement mode
    instrument.write(":PAGE:MEAS:CONS:SMU1 0")  # sets the constant SOURCE value of SMU1 for sweep measurements
    instrument.write(":PAGE:MEAS:CONS:SMU2 0")  # sets the constant SOURCE value of SMU2 for sweep measurements
    instrument.write(":PAGE:MEAS:CONS:SMU3 0")  # sets the constant SOURCE value of SMU2 for sweep measurements
    instrument.write(":PAGE:MEAS:CONS:SMU4 0")  # sets the constant SOURCE value of SMU2 for sweep measurements
    instrument.write(":PAGE:MEAS:CONS:SMU2:COMP 0.1")

def get_measurements():
    instrument.write(":PAGE:GLIS:SCAL:AUTO ONCE")
    instrument.write(":PAGE:SCON:SING")  # starts the single measurement operation
    instrument.query("*OPC?")  # starts to monitor pending operations, and sets/clears the Operation complete
    instrument.write('FORM:DATA ASC')  # specifies the data format as ASCII
    v_4 = instrument.query_ascii_values(":DATA? 'V4'")  # gets the data of V3 measurements
    i_1 = instrument.query_ascii_values(":DATA? 'I1'")  # gets the data of I1 measurements
    i_2 = instrument.query_ascii_values(":DATA? 'I2'")  # gets the data of I1 measurements
    i_3 = instrument.query_ascii_values(":DATA? 'I3'")  # gets the data of I1 measurements
    i_4 = instrument.query_ascii_values(":DATA? 'I4'")  # gets the data of I3 measurements
    return v_4, i_1, i_2, i_3, i_4


def get_minimal_measures():
    print('wait_for_stable')

    linear_fun = lambda x, a, b: a + b * x
    sets_of_measurements = []
    while True:
        v_4, i_1, i_2, i_3, i_4 = get_measurements()
        popt, pcov = curve_fit(f=linear_fun, xdata=v_4, ydata=i_4, p0=[0, 0], bounds=(-np.inf, np.inf))
        # print('popt::', popt, '; pcov::', pcov)
        # print('zero::', i_4[10])
        Voffset = -popt[0] / popt[1]
        # print('Voffset:', Voffset)
        sets_of_measurements.append({'i_1': i_1, 'i_2': i_2, 'i_4': i_4, 'v_4': v_4, 'i_3': i_3, 'offset': abs(popt[0])})
        if len(sets_of_measurements) >= 2:
            if sets_of_measurements[-1]['offset'] > sets_of_measurements[-2]['offset']:
                return v_4, i_1, i_2, i_3, i_4

        sleep(1)


def fill_condition_list(sheet: Worksheet, configs: dict):
    pattern = re.compile('%([\\w\\d_]+)%', re.I)
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == None:
                continue
            match = pattern.match(cell.value)
            if match is not None and match.group(1) in configs:
                cell.value = configs[match.group(1)]


def fill_data_list(sheet: Worksheet, data: dict):
    indexes = list(string.ascii_uppercase)
    for column_name in data:
        letter_index = indexes.pop(0)
        sheet["{}1".format(letter_index)] = column_name
        values = data[column_name]
        for i in range(len(values)):
            sheet["{}{}".format(letter_index, i + 2)] = values[i]

def save_data_to_bd(configs: dict, V3: list, I1: list, I3: list, chip_name: str):
    db_configs = {
    'user': 'prober',
    'password': 'eU1vf0k7L916#',
    'host': '95.217.222.91',
    'database': 'elfys',
    'raise_on_warnings': True
    }
    try:
        cnx = mysql.connector.connect(**db_configs)
        cursor = cnx.cursor()
        query = """
        INSERT INTO iv_data (wafer, chip, int_time, temperature, voltage_input, anode_current, cathode_current, anode_current_corrected, stage)
        VALUES (%(wafer)s, %(chip)s, %(int_time)s, %(temperature)s, %(voltage_input)s,  %(anode_current)s, %(cathode_current)s, %(anode_current_corrected)s, %(stage)s)
        """
        for voltage_input, anode_current, cathode_current in zip(V3, I1, I3):
            anode_current_corrected = compute_corrected_current(
                configs['TEMPERATURE'], anode_current)
            cursor.execute(
                query, {'wafer': WAFER, 'chip': chip_1, 'int_time': configs['INT_TIME'], 'temperature': configs['TEMPERATURE'], 'voltage_input': voltage_input, 'anode_current': anode_current, 'cathode_current': cathode_current, 'anode_current_corrected': anode_current_corrected, 'stage': stage})
            cnx.commit()
            cursor.execute(
                query, {'wafer': WAFER, 'chip': chip_2, 'int_time': configs['INT_TIME'], 'temperature': configs['TEMPERATURE'], 'voltage_input': voltage_input, 'anode_current': anode_current, 'cathode_current': cathode_current, 'anode_current_corrected': anode_current_corrected, 'stage': stage})
            cnx.commit()
            cursor.execute(
                query, {'wafer': WAFER, 'chip': chip_3, 'int_time': configs['INT_TIME'], 'temperature': configs['TEMPERATURE'], 'voltage_input': voltage_input, 'anode_current': anode_current, 'cathode_current': cathode_current, 'anode_current_corrected': anode_current_corrected, 'stage': stage})
            cnx.commit()
    except mysql.connector.Error as err:
        if err.errno == errorcode.ER_ACCESS_DENIED_ERROR:
            print("Something is wrong with your user name or password")
        elif err.errno == errorcode.ER_BAD_DB_ERROR:
            print("Database does not exist")
        else:
            print(err)
    finally:
        if cnx.is_connected():
            cursor.close()
            cnx.close()



def compute_corrected_current(temp: float, current: float):
    return 1.15 ** (TARGET_TEMPERATURE - temp) * current #1e+12 added in order to see numbers in the table, otherwise there are zeros


set_configs(configs)
data_v4, data_i1, data_i2, data_i3, data_i4 = get_minimal_measures()

save_data_to_bd(configs, data_v4, data_i1, data_i3, chip_1)
save_data_to_bd(configs, data_v4, data_i1, data_i3, chip_2)
save_data_to_bd(configs, data_v4, data_i1, data_i3, chip_3)


configs_20v = {
    "VAR1_START": 20,
    "VAR1_STEP": -0.5,  # step voltage
    "VAR1_END": -2,  # end voltage
    "C_VAR1": 0.1,  # compliance
    "HOLD_TIME": 3,
    "DELAY_TIME": 0,
    "INT_TIME": 'MED',
    "INT_TIME_SHORT": '',
    "INT_TIME_MED": '',
    "INT_TIME_LONG": '',
    "VAR1_MODE": 'SINGLE',
    "VAR1_SWEEP_TYPE": 'LIN',
    "VOLT_OUTPUT_SMU1": 'V',
    "VOLT_OUTPUT_SMU2": 'V',
    "VOLT_OUTPUT_SMU3": 'V',
    "VOLT_OUTPUT_SMU4": 'V',
    "FUNCTION_SMU1": 'CONS',
    "FUNCTION_SMU2": 'CONS',
    "FUNCTION_SMU3": 'CONS',
    "FUNCTION_SMU4": 'VAR1',
    "V_NAME_SMU1": 'V1',
    "V_NAME_SMU2": 'V2',
    "V_NAME_SMU3": 'V3',
    "V_NAME_SMU4": 'V4',
    "I_NAME_SMU1": 'I1',
    "I_NAME_SMU2": 'I2',
    "I_NAME_SMU3": 'I3',
    "I_NAME_SMU4": 'I4',
    "RANGE_MODE_SMU1": 'LIM',
    "RANGE_MODE_SMU2": 'LIM',
    "RANGE_MODE_SMU3": 'LIM',
    "RANGE_MODE_SMU4": 'LIM',
    "RANGE_VALUE_SMU1": '1E-9',
    "RANGE_VALUE_SMU2": '1E-9',
    "RANGE_VALUE_SMU3": '1E-9',
    "RANGE_VALUE_SMU4": '1E-9',
    "DATETIME": datetime.datetime.now(),
    "TEMPERATURE": get_temperature()
}


def sweep():
    instrument.write(":PAGE:CHAN:MODE SWEEP")  # sets the sweep measurement mode
    instrument.write(":PAGE:SCON:SING")  # starts the single measurement operation

    instrument.query("*OPC?")  # starts to monitor pending operations, and sets/clears the Operation complete
    instrument.write('FORM:DATA ASC')  # specifies the data format as ASCII
    v_4 = instrument.query_ascii_values(":DATA? 'V4'")  # gets the data of I1 measurements
    i_1 = instrument.query_ascii_values(":DATA? 'I1'")
    i_2 = instrument.query_ascii_values(":DATA? 'I2'")  # gets the data of I1 measurements
    i_3 = instrument.query_ascii_values(":DATA? 'I3'")  # gets the data of I3 measurements
    i_4 = instrument.query_ascii_values(":DATA? 'I4'")
    return v_4, i_1, i_2, i_3, i_4


def save_20v_measurements(configs_20v: dict, V3: list, I1: list, I3: list, chip_name: str):
    wb = load_workbook(filename="template_linear1.xlsx")
    fill_condition_list(wb['conditions'], configs_20v)
    corrector = compute_corrected_current(configs_20v['TEMPERATURE'])
    fill_data_list(wb['data'], {
        'V3': V3,
        'I1': I1,
        'I3': I3,
        'corrected value': list(map(corrector, I1))
    })
    filepath = 'IV_measurements\\20V\\{}.xlsx'.format(chip_name)
    wb.save(filepath)
    wb.close()


set_configs(configs_20v)
data_v4, data_i1, data_i2, data_i3, data_i4 = sweep()

save_data_to_bd(configs_20v, data_v4, data_i1, data_i3, chip_1)
save_data_to_bd(configs_20v, data_v4, data_i1, data_i3, chip_2)
save_data_to_bd(configs_20v, data_v4, data_i1, data_i3, chip_3)

try:
    rm.close()
except:
    ...
