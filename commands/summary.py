from decimal import Decimal
from os.path import exists as file_exists
from time import strftime, localtime
from typing import Union, Any

import click
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from matplotlib.ticker import MaxNLocator
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
from sqlalchemy.orm import Session, joinedload

from orm import IVMeasurement, Wafer, Chip
from utils import logger


@click.command(name='summary', help="Make summary files (png and xlsx) for IV measurements' data.")
@click.pass_context
@click.option("-t", "--chips-type", help="Type of the chips to analyze.")
@click.option("-w", "--wafer", "wafer_name", prompt=f"Wafer name", help="Wafer name.")
@click.option("-o", "--output", "file_name", default=lambda: f"summary-{strftime('%y%m%d-%H%M%S')}",
              help="Output file names without extension.", show_default="summary-{datetime}")
@click.option("-s", "--chip-state", "chip_states", help="State of the chips to analyze.",
              default=['all'], show_default=True, multiple=True)
@click.option("--outliers-coefficient", default=2.0, show_default=True,
              help="Standard deviation multiplier to detect outlier measurements.", type=float)
def summary(ctx: click.Context, chips_type: Union[str, None], wafer_name: str, file_name: str,
            chip_states: list[str],
            outliers_coefficient: float):
    session: Session = ctx.obj['session']
    if ctx.obj['default_wafer'].name != wafer_name:
        wafer = session.query(Wafer).filter(Wafer.name == wafer_name).first()
    else:
        wafer = ctx.obj['default_wafer']
    query = session.query(IVMeasurement) \
        .filter(IVMeasurement.chip.has(Chip.wafer.__eq__(wafer))) \
        .options(joinedload(IVMeasurement.chip))

    if chips_type is not None:
        query = query.filter(IVMeasurement.chip.has(Chip.type.__eq__(chips_type)))
    else:
        logger.info('No chips type specified. Analyzing all chips.')

    if 'all' not in chip_states:
        query = query.filter(IVMeasurement.chip_state_id.in_(chip_states))

    measurements = query.all()

    if not measurements:
        logger.error('No measurements found.')
        return

    sheets_data = get_sheets_data(measurements)
    plot_summary_voltages = list(map(Decimal, ["0.01", "5"]))
    fig = plot_data(measurements, plot_summary_voltages, outliers_coefficient=outliers_coefficient)

    png_file_name = file_name + '.png'
    check_file_exists(png_file_name)
    fig.savefig(png_file_name, dpi=300)
    logger.info(f'Summary data is plotted to {png_file_name}')

    exel_file_name = file_name + '.xlsx'
    check_file_exists(exel_file_name)
    info = get_info(ctx, wafer=wafer, chip_states=chip_states)
    save_summary_to_excel(sheets_data, info, exel_file_name)

    logger.info(f'Summary data is saved to {exel_file_name}')


def save_summary_to_excel(sheets_data: dict[str, pd.DataFrame], info: pd.Series, file_name: str):
    with pd.ExcelWriter(file_name) as writer:
        summary_voltages = list(map(Decimal, ["-1", "0.01", "5", "10", "20"]))
        summary_df = sheets_data['anode'][summary_voltages].rename(columns=float)
        summary_df.to_excel(writer, sheet_name='Summary')
        summary_sheet = writer.book["Summary"]
        red_fill = PatternFill(bgColor='ee9090', fill_type='solid')
        green_fill = PatternFill(bgColor='90ee90', fill_type='solid')

        chips_row_numbers = [(i + 2, name) for i, name in enumerate(sheets_data['chip_names'])]

        for chip_type in sheets_data['chip_types']:

            def is_current_type(chip_name: str) -> bool:
                return chip_name.startswith(chip_type)

            if chip_type == "X":
                thresholds = {'-1': 1e-3, '0.01': -12e-12, '10': -80e-12, '20': -500e-12}
            elif chip_type == 'Y':
                thresholds = {'-1': 1e-3, '0.01': -15e-12, '10': -120e-12, '20': -1000e-12}
            elif chip_type == 'U':
                thresholds = {'-1': 1e-3, '0.01': -30e-12, '10': -200e-12, '20': -1200e-12}
            elif chip_type == 'V':
                thresholds = {'-1': 1e-3, '0.01': -60e-12, '10': -800e-12, '20': -1400e-12}
            elif chip_type == 'F':
                thresholds = {'-1': 5e-3, '0.01': -40e-12, '10': -2e-9, }
            elif chip_type == 'G':
                thresholds = {'-1': 5e-3, '0.01': -50e-12, '10': -2e-9, }
            elif chip_type == 'E':
                thresholds = {'-1': 5e-3, '0.01': -20e-12, '10': -2e-9, }
            elif chip_type == 'C':
                thresholds = {'-1': 1e-3, '0.01': -20e-12, '10': -1e-9, }
            else:
                continue

            for voltage, current_threshold in thresholds.items():
                try:
                    column_index = summary_voltages.index(Decimal(voltage))

                    first_row_index = next(i for i, v in chips_row_numbers if is_current_type(v))
                    last_row_index = next(
                        i for i, v in reversed(chips_row_numbers) if is_current_type(v))
                except ValueError:
                    continue

                column_letter = chr(ord('B') + column_index)
                cell_range = f'{column_letter}{first_row_index}:{column_letter}{last_row_index}'
                summary_sheet.conditional_formatting \
                    .add(cell_range, CellIsRule(operator='lessThan', formula=[current_threshold],
                                                fill=red_fill))
                summary_sheet.conditional_formatting \
                    .add(cell_range,
                         CellIsRule(operator='greaterThanOrEqual', formula=[current_threshold],
                                    fill=green_fill))

        # summary_sheet.

        sheets_data['anode'].rename(columns=float).to_excel(writer, sheet_name='I1 anode')
        sheets_data['cathode'].rename(columns=float).to_excel(writer, sheet_name='I3 cathode')
        info.to_excel(writer, sheet_name='Info')


def get_sheets_data(measurements: list[IVMeasurement]) -> dict[str, Union[pd.DataFrame, Any]]:
    chip_names = sorted({measurement.chip.name for measurement in measurements})
    chip_types = sorted({measurement.chip.type for measurement in measurements})
    voltages = {measurement.voltage_input for measurement in measurements}
    anode_df = pd.DataFrame(dtype='float64', index=chip_names, columns=voltages)
    cathode_df = pd.DataFrame(dtype='float64', index=chip_names, columns=voltages)
    with click.progressbar(measurements, label='Processing measurements...') as progress:
        for measurement in progress:
            measurement: IVMeasurement
            anode_df.loc[
                measurement.chip.name, measurement.voltage_input] = measurement.anode_current_corrected
            cathode_df.loc[
                measurement.chip.name, measurement.voltage_input] = measurement.cathode_current
    return {
        'anode': anode_df,
        'cathode': cathode_df,
        'chip_names': chip_names,
        'chip_types': chip_types,
        'voltages': voltages
    }


def get_info(ctx: click.Context, wafer: Wafer, chip_states: list[str]) -> pd.Series:
    format_date = strftime("%A, %d %b %Y", localtime())
    if 'all' in chip_states:
        chip_states_str = 'all'
    else:
        chip_states_str = "; ".join(
            [state.name for state in ctx.obj['chip_states'] if str(state.id) in chip_states])

    return pd.Series({
        'Wafer': wafer.name,
        'Summary generation date': format_date,
        'Chip state': chip_states_str,
    })


def check_file_exists(file_name: str):
    while True:
        if file_exists(file_name):
            ok = click.prompt(f'File {file_name} already exists. Overwrite?', default='y',
                              type=click.Choice(['y', 'n']))
            if ok == 'n':
                file_name = click.prompt(f'Enter new file name')
            else:
                break
        else:
            break


def get_heat_map_data(measurements: list[IVMeasurement, ...]) -> dict:
    xs, ys = [], []

    for measurement in measurements:
        xs.append(measurement.chip.x_coordinate)
        ys.append(measurement.chip.y_coordinate)

    borders = {'left': min(xs), 'right': max(xs), 'bottom': min(ys), 'top': max(ys)}
    width = borders['right'] - borders['left'] + 1
    height = borders['top'] - borders['bottom'] + 1
    data = np.full((height, width), np.nan)
    for measurement in measurements:
        data[
            measurement.chip.y_coordinate - borders['bottom']][
            measurement.chip.x_coordinate - borders['left']] = measurement.anode_current_corrected
    return {'data': data, 'borders': borders}


def get_outliers_idx(data: np.ndarray, m: float) -> np.ndarray:
    return np.abs(data - np.nanmedian(data)) > m * np.nanstd(data)


def plot_hist(ax: Axes, data: np.ndarray):
    ax.set_ylabel("Number of chips")
    ax.set_xlabel("Anode current [pA]")
    ax.hist(data * 1e12, bins=15)


def plot_heat_map(ax: Axes, measurements: list[IVMeasurement, ...], low, high):
    xs = {measurement.chip.x_coordinate for measurement in measurements}
    ys = {measurement.chip.y_coordinate for measurement in measurements}

    width = max(xs) - min(xs) + 1
    height = max(ys) - min(ys) + 1
    grid = np.full((height, width), np.nan)
    for cell in measurements:
        grid[cell.chip.y_coordinate - min(ys)][
            cell.chip.x_coordinate - min(xs)] = cell.anode_current_corrected

    X = np.linspace(min(xs) - 0.5, max(xs) + 0.5, width + 1)
    Y = np.linspace(min(ys) - 0.5, max(ys) + 0.5, height + 1)
    mesh = ax.pcolormesh(X, Y, grid, cmap='hot', shading='flat', vmin=low, vmax=high)
    ax.xaxis.set_major_locator(MaxNLocator(integer=True, min_n_ticks=0))
    ax.yaxis.set_major_locator(MaxNLocator(integer=True, min_n_ticks=0))
    ax.set_ylabel("Y coordinate")
    ax.set_xlabel("X coordinate")
    ax.figure.colorbar(mesh, ax=ax)


def plot_data(measurements: list[IVMeasurement, ...],
              summary_voltages: list[Decimal, ...], outliers_coefficient: float) -> Figure:
    fig, axes = plt.subplots(nrows=len(summary_voltages), ncols=2,
                             figsize=(10, 5 * len(summary_voltages)),
                             gridspec_kw=dict(left=0.08, right=0.95, bottom=0.05, top=0.95,
                                              wspace=0.3, hspace=0.35))
    for i, voltage in enumerate(summary_voltages):
        target_measurements = [measurement for measurement in measurements if
                               measurement.voltage_input == voltage]
        data = np.array(
            [measurement.anode_current_corrected for measurement in target_measurements])

        outliers_idx = get_outliers_idx(data, outliers_coefficient)
        if outliers_idx.any():
            outlier_chip_names = (outlier.chip.name for outlier in
                                  np.array(target_measurements)[outliers_idx])
            logger.warn(
                f'Outliers detected! {", ".join(outlier_chip_names)} are ignored on {voltage}V histogram and heat map color scale')
            data = data[~outliers_idx]

        axes[i][0].set_title(f"{voltage}V")
        plot_hist(axes[i][0], data)

        low, high = data.min(), data.max()
        axes[i][1].set_title(f"{voltage}V")
        plot_heat_map(axes[i][1], target_measurements, low, high)
    return fig
